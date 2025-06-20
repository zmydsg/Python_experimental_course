# Module for Python course project V3.0 2025
# This module provide a competition for player functions
# The player functions are saved in .py files in the  
#   folder "player" under the same path with this file
# The player functions must have a function (x,y) = player(Colour, Board)
from Reversi import *
import random, os, sys
import importlib
import time
import datetime
from math import log

from openpyxl import Workbook, load_workbook

# 创建日志文件
log_filename = f"game_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
log_file = open(log_filename, 'w')

old_stdout = sys.stdout

# 同时输出到控制台和日志文件的函数
def log_print(*args, **kwargs):
    # 保存原始的end和sep值
    end_val = kwargs.get('end', '\n')
    sep_val = kwargs.get('sep', ' ')
    
    # 打印到控制台
    print(*args, **kwargs)
    
    # 打印到日志文件
    print(*args, sep=sep_val, end=end_val, file=log_file)
    log_file.flush()  # 确保立即写入文件

# Disable
def blockPrint():
    sys.stdout = open(os.devnull, 'w')

# Restore
def enablePrint():
    sys.stdout = old_stdout

# path and files
sys.path.append(sys.path[0]+'\\players')
FileNames = [x for x in os.listdir('players') if x.endswith(".py")]
StudentList = [x.strip('.py') for x in os.listdir('players')
               if x.endswith(".py") and x.startswith('D') and len(x) == 15]

wb = Workbook()
ws = wb.active
StartTime = datetime.datetime.now()

# 记录游戏开始时间到日志文件
log_print(f"游戏开始时间: {StartTime.strftime('%Y-%m-%d %H:%M:%S')}")

play_round = 1#100
pk_function = []
pk_player = []
pk_time = []
for k in range(len(StudentList)):
    play = importlib.util.find_spec(StudentList[k])
    if (play is not None):
        try:
            log_print(f"loading player function from {StudentList[k]}")
            play = importlib.import_module(StudentList[k])
            pk_function.append(play.player)
            pk_player.append(StudentList[k])
            pk_time.append(0)
        except Exception as e:
            log_print(f"exception happen when loading player function from {StudentList[k]}\                \r\n the exception is {e}")
            buf = [str(e)]
            ws.append(buf)
log_print(f"Loaded players are: {pk_player}", f"Total number of players is: {len(pk_player)}")
buf = ['student']
buf.extend(pk_player)
ws.append(buf)

buf = []
buf1 = []
mark = []
timecount = []
illegal_player = []
slow_player = []
error_player = []
error_message = []

for i in range(len(pk_function)):
    buf.append([])
    mark.append([])
    for j in range(len(pk_function)):
        mark[i].append(0)
for i in range(len(pk_function)):
    buf[i].append(pk_player[i])
    for j in range(i+1, len(pk_function)):
        if illegal_player.count(i) > 0:
            break
        if slow_player.count(i) >0:
            break
        if error_player.count(i) >0:
            break
        start = time.time()
        for round in range(play_round):
            if illegal_player.count(j) > 0:
                break
            if slow_player.count(j) >0:
                break
            if error_player.count(j) >0:
                break
            blockPrint()
            Board, result, PlayTime, error = PlayGame(pk_function[j], pk_function[i])
            enablePrint()
            pk_time[j] = pk_time[j] + PlayTime[0]
            pk_time[i] = pk_time[i] + PlayTime[1]
            if result == PlayerSlow:
                log_print(pk_player[j], 'is a slow player')
                slow_player.append(j)
                mark[i][j] = 0
                break
            elif result == PlayerError:
                log_print(pk_player[j], 'has error', error)
                error_player.append(j)
                error_message.append(error)
                mark[i][j] = 0
                break
            elif result == IllegalMove:
                log_print(pk_player[j], 'illegal move')
                illegal_player.append(j)
                mark[i][j] = 0
                break
            elif result == -PlayerSlow:
                log_print(pk_player[i], 'is a slow player')
                slow_player.append(i)
                mark[j][i] = 0
                break
            elif result == -IllegalMove:
                log_print(pk_player[i], 'illegal move')
                illegal_player.append(i)
                mark[j][i] = 0
                break
            elif result == -PlayerError:
                log_print(pk_player[i], 'has error', error)
                error_player.append(i)
                error_message.append(error)
                mark[j][i] = 0
                break
            elif result > 0:
                mark[i][j] = mark[i][j] + 3
            elif result < 0:
                mark[j][i] = mark[j][i] + 3
            elif result == 0:
                mark[i][j] = mark[i][j] + 1
                mark[j][i] = mark[j][i] + 1
        for round in range(play_round):
            if illegal_player.count(j) != 0 or slow_player.count(j) != 0 \
               or error_player.count(j) != 0:
                break
            blockPrint()
            Board, result, PlayTime, error = PlayGame(pk_function[i], pk_function[j])
            enablePrint()
            pk_time[j] = pk_time[j] + PlayTime[1]
            pk_time[i] = pk_time[i] + PlayTime[0]
            if result == PlayerSlow:
                log_print(pk_player[i], 'is a slow player')
                slow_player.append(i)
                mark[j][i] = 0
                break
            elif result == IllegalMove:
                log_print(pk_player[i], 'illegal move')
                illegal_player.append(i)
                mark[j][i] = 0
                break
            elif result == PlayerError:
                log_print(pk_player[i], 'has error', error)
                error_player.append(i)
                error_message.append(error)
                mark[j][i] = 0
                break
            elif result == -PlayerSlow:
                log_print(pk_player[j], 'is a slow player')
                slow_player.append(j)
                mark[i][j] = 0
                break
            elif result == -IllegalMove:
                log_print(pk_player[j], 'illegal move')
                illegal_player.append(j)
                mark[i][j] = 0
                break
            elif result == -PlayerError:
                log_print(pk_player[j], 'has error', error)
                error_player.append(j)
                error_message.append(error)
                mark[i][j] = 0
                break
            elif result < 0:
                mark[i][j] = mark[i][j] + 3
            elif result > 0:
                mark[j][i] = mark[j][i] + 3
            elif result == 0:
                mark[i][j] = mark[i][j] + 1
                mark[j][i] = mark[j][i] + 1
        log_print(pk_player[i]+'('+ "%.5f" %  pk_time[i], 's) vs ',
              pk_player[j]+'('+ '%.5f' %  pk_time[j], 's)',
              #'%.5f' % (time.time()-start),
              mark[i][j], ':', mark[j][i])
        
for i in (error_player + illegal_player + slow_player):
    for j in range(len(pk_function)):
        mark[i][j] = 0
        mark[j][i] = 0    

for i in range(len(pk_function)):
    buf[i].extend(mark[i])
    ws.append(buf[i])
try:
    wb.save("sample.xlsx")
    log_print("sample.xlsx is saved")
except:
    wb.save("sample1.xlsx")
    log_print("sample1.xlsx is saved")

TotalTime = sum(pk_time)
log_print('Total Time PK time is:', '%.2f' % TotalTime)

pk_time_average = sum(pk_time)/len(pk_time)
pk_time_max = max(pk_time)
pk_time_min = min(pk_time)
EndTime = datetime.datetime.now()
TimeUsed = EndTime - StartTime
TimeUsedSec = TimeUsed.seconds
TimeUsedHour = TimeUsedSec // 3600
TimeUsedSec = TimeUsedSec % 3600
TimeUsedMin = TimeUsedSec // 60
TimeUsedSec = TimeUsedSec % 60
TimeStamp = 'Started at: ' + str(StartTime)[0:19] + '; ' + str(play_round) + \
            ' trails; Time consumed: %02.0f' % TimeUsedHour + ':' + \
            '%02.0f' % TimeUsedMin + ':' + '%03.1f' % TimeUsedSec
log_print(TimeStamp)

# 记录游戏结束时间到日志文件
log_print(f"游戏结束时间: {EndTime.strftime('%Y-%m-%d %H:%M:%S')}")

wb = Workbook()
ws = wb.active
#ws.append([TimeStamp, len(pk_function)])
ws.append(['Student', 'Time Percentage', 'Time Used', 'Time Rank', 'Factor', 'Score',
           'Score Rank', 'Final', 'Final Rank', 'Remark'])

for i in range(len(pk_function)):
    Remark = ' '
    log_print(pk_player[i], " time percetage:", '%.3f' % (pk_time[i]/TotalTime*100), " time:", '%.5f' % pk_time[i])
    score = sum(buf[i][1:len(buf[i])])
    if illegal_player.count(i) > 0 or slow_player.count(i) > 0 \
       or error_player.count(i) >0 :
        factor = 1
        score = 0
    else:
        factor = log(pk_time_average/pk_time[i]+10, 10)
    final = score * factor
    buf1 = [pk_player[i], str(pk_time[i]/TotalTime*100), pk_time[i], '=rank(c' + str(i+2) + ', c2:c80,1)',
            factor, score, '=rank(f' + str(i+2) + ', f2:f80)', final, '=rank(h' + str(i+2) + ', h2:h80)']
    if illegal_player.count(i) > 0:
        Remark = Remark + 'Illegal Move '
    elif slow_player.count(i) > 0:
        Remark = Remark + 'Slow Player '
    elif error_player.count(i) > 0:
        Remark = Remark + 'Error: ' + str(error_message[error_player.index(i)])
    buf1.append(Remark)
    ws.append(buf1)
buf = []
try:
    wb.save("time.xlsx")
    log_print("result saved in time.xlsx")
except:
    wb.save("time1.xlsx")
    log_print("result saved in time1.xlsx")

# 关闭日志文件
log_print(f"日志文件已保存为: {log_filename}")
log_file.close()
